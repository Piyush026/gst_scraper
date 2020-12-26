import openpyxl

path = "Company and TAN.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
# print(m_row)
lst = []
for i in range(50, 100):
    cell_obj = sheet_obj.cell(row=i, column=1)
    lst.append(cell_obj.value)
print(lst)
